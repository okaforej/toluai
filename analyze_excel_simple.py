#!/usr/bin/env python3
import sys
sys.path.append('/Users/eokafor/Library/Python/3.9/lib/python/site-packages')

from openpyxl import load_workbook
import os
import json

# Load the workbook
file_path = os.path.expanduser('~/Downloads/IRPA CCI Model.xlsx')
wb = load_workbook(file_path, read_only=True, data_only=True)

print('=== EXCEL SHEETS FOUND ===')
for sheet_name in wb.sheetnames:
    print(f'  - {sheet_name}')
print()

# Store all data
all_data = {}

# Read each sheet
for sheet_name in wb.sheetnames:
    print(f'\n=== SHEET: {sheet_name} ===')
    sheet = wb[sheet_name]
    
    # Get data from sheet
    data = []
    headers = None
    row_count = 0
    
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        # Skip completely empty rows
        if all(cell is None for cell in row):
            continue
            
        if row_idx == 0:
            # First non-empty row is headers
            headers = [str(cell) if cell else f'Column_{i}' for i, cell in enumerate(row)]
            print(f'Columns: {headers}')
        else:
            # Create dict from row data
            row_dict = {}
            for i, header in enumerate(headers):
                if i < len(row) and row[i] is not None:
                    row_dict[header] = row[i]
            if row_dict:  # Only add non-empty rows
                data.append(row_dict)
                row_count += 1
    
    all_data[sheet_name] = data
    print(f'Total rows with data: {row_count}')
    
    # Show first few rows
    if data:
        print('\nFirst 5 rows:')
        for i, row_data in enumerate(data[:5], 1):
            print(f'\nRow {i}:')
            for key, value in row_data.items():
                print(f'  {key}: {value}')
        
        if len(data) > 5:
            print(f'\n... and {len(data) - 5} more rows')
    
    print('=' * 80)

# Close workbook
wb.close()

# Save to JSON
output_file = '/Users/eokafor/toluai/irpa_specs.json'
with open(output_file, 'w') as f:
    json.dump(all_data, f, indent=2, default=str)
    
print(f'\n✅ Data saved to {output_file}')
print(f'Total sheets processed: {len(all_data)}')
for sheet_name, data in all_data.items():
    print(f'  - {sheet_name}: {len(data)} rows')